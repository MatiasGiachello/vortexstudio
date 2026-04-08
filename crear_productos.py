from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# =============================================
# PRODUCTO 3: PACK PRESUPUESTOS PRO (Excel)
# =============================================
wb = Workbook()

# ---- HOJA 1: PRESUPUESTO ----
ws1 = wb.active
ws1.title = "Presupuesto"

# Colores
azul_oscuro = "0F3A9E"
azul_medio = "1A56DB"
azul_claro = "E8F0FE"
gris_claro = "F7F8FA"
gris_borde = "D0D5DD"
blanco = "FFFFFF"
verde = "22C55E"
texto_oscuro = "0D0F14"
texto_gris = "5A6270"

def header_font(bold=True, size=11, color="FFFFFF"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color="0D0F14"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def fill(color):
    return PatternFill("solid", start_color=color, end_color=color)

def border(style="thin", color="D0D5DD"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# Anchos de columna
ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 32
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 14
ws1.column_dimensions["F"].width = 16

# Altura de filas clave
for r in range(1, 60):
    ws1.row_dimensions[r].height = 18

ws1.row_dimensions[1].height = 8
ws1.row_dimensions[2].height = 50
ws1.row_dimensions[3].height = 22

# --- HEADER ---
ws1.merge_cells("B2:D2")
ws1["B2"] = "VORTEX STUDIO"
ws1["B2"].font = Font(name="Arial", bold=True, size=22, color=blanco)
ws1["B2"].fill = fill(azul_oscuro)
ws1["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

ws1.merge_cells("E2:F2")
ws1["E2"] = "PRESUPUESTO"
ws1["E2"].font = Font(name="Arial", bold=True, size=16, color=blanco)
ws1["E2"].fill = fill(azul_medio)
ws1["E2"].alignment = center()

# --- INFO EMPRESA / CLIENTE ---
ws1.merge_cells("B3:D3")
ws1["B3"] = "Agencia de presencia digital · hola@vortexstudio.com"
ws1["B3"].font = body_font(size=9, color=texto_gris)
ws1["B3"].fill = fill(azul_claro)
ws1["B3"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

ws1["E3"] = "N° Presupuesto:"
ws1["E3"].font = body_font(bold=True, size=9)
ws1["E3"].alignment = Alignment(horizontal="right", vertical="center")

ws1["F3"] = "=\"PRES-\"&TEXT(TODAY(),\"YYYYMMDD\")"
ws1["F3"].font = body_font(bold=True, size=9, color=azul_oscuro)
ws1["F3"].alignment = center()

# --- DATOS CLIENTE ---
ws1.row_dimensions[5].height = 20
ws1.row_dimensions[6].height = 20
ws1.row_dimensions[7].height = 20
ws1.row_dimensions[8].height = 20
ws1.row_dimensions[9].height = 20

ws1.merge_cells("B5:C5")
ws1["B5"] = "DATOS DEL CLIENTE"
ws1["B5"].font = header_font(size=9, color=azul_oscuro)
ws1["B5"].fill = fill(azul_claro)
ws1["B5"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws1.merge_cells("E5:F5")
ws1["E5"] = "DETALLES DEL PRESUPUESTO"
ws1["E5"].font = header_font(size=9, color=azul_oscuro)
ws1["E5"].fill = fill(azul_claro)
ws1["E5"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

campos_cliente = [
    ("B6", "Empresa / Nombre:", "C6", ""),
    ("B7", "Email:", "C7", ""),
    ("B8", "Teléfono:", "C8", ""),
    ("B9", "Dirección:", "C9", ""),
]
campos_detalle = [
    ("E6", "Fecha:", "F6", "=TODAY()"),
    ("E7", "Válido hasta:", "F7", "=TODAY()+30"),
    ("E8", "Moneda:", "F8", "USD"),
    ("E9", "Forma de pago:", "F9", "50% adelanto / 50% entrega"),
]

for label_cell, label_val, val_cell, val_val in campos_cliente:
    ws1[label_cell] = label_val
    ws1[label_cell].font = body_font(bold=True, size=9, color=texto_gris)
    ws1[label_cell].alignment = left()
    ws1[val_cell] = val_val
    ws1[val_cell].font = body_font(size=9)
    ws1[val_cell].alignment = left()
    ws1[val_cell].border = Border(bottom=Side(style="thin", color=gris_borde))

for label_cell, label_val, val_cell, val_val in campos_detalle:
    ws1[label_cell] = label_val
    ws1[label_cell].font = body_font(bold=True, size=9, color=texto_gris)
    ws1[label_cell].alignment = Alignment(horizontal="right", vertical="center")
    ws1[val_cell] = val_val
    ws1[val_cell].font = body_font(bold=True, size=9, color=azul_oscuro)
    ws1[val_cell].alignment = center()
    if val_val == "=TODAY()" or val_val == "=TODAY()+30":
        ws1[val_cell].number_format = "DD/MM/YYYY"

# --- TABLA DE ITEMS ---
ws1.row_dimensions[11].height = 22
headers = ["#", "Descripción del servicio / producto", "Cantidad", "Precio unit.", "Descuento", "Subtotal"]
cols = ["A", "B", "C", "D", "E", "F"]
for i, (col, h) in enumerate(zip(cols, headers)):
    c = ws1[f"{col}11"]
    c.value = h
    c.font = header_font(size=9)
    c.fill = fill(azul_oscuro)
    c.alignment = center()
    c.border = border(color=azul_medio)

# Items de ejemplo
items = [
    (1, "Diseño web profesional (landing page)", 1, 150, 0),
    (2, "Identidad de marca (logo + paleta + guía)", 1, 120, 0),
    (3, "Gestión de redes sociales (mensual)", 3, 80, 10),
    (4, "", 0, 0, 0),
    (5, "", 0, 0, 0),
    (6, "", 0, 0, 0),
    (7, "", 0, 0, 0),
    (8, "", 0, 0, 0),
]

for idx, (num, desc, qty, price, disc) in enumerate(items):
    row = 12 + idx
    ws1.row_dimensions[row].height = 20
    bg = blanco if idx % 2 == 0 else gris_claro

    ws1[f"A{row}"] = num if num else ""
    ws1[f"A{row}"].font = body_font(size=9, color=texto_gris)
    ws1[f"A{row}"].fill = fill(bg)
    ws1[f"A{row}"].alignment = center()

    ws1[f"B{row}"] = desc
    ws1[f"B{row}"].font = body_font(size=9)
    ws1[f"B{row}"].fill = fill(bg)
    ws1[f"B{row}"].alignment = left()

    ws1[f"C{row}"] = qty if qty else ""
    ws1[f"C{row}"].font = body_font(size=9)
    ws1[f"C{row}"].fill = fill(bg)
    ws1[f"C{row}"].alignment = center()

    ws1[f"D{row}"] = price if price else ""
    ws1[f"D{row}"].font = body_font(size=9)
    ws1[f"D{row}"].fill = fill(bg)
    ws1[f"D{row}"].alignment = center()
    ws1[f"D{row}"].number_format = "$#,##0.00"

    ws1[f"E{row}"] = f"{disc}%" if disc else ""
    ws1[f"E{row}"].font = body_font(size=9, color=verde if disc else texto_oscuro)
    ws1[f"E{row}"].fill = fill(bg)
    ws1[f"E{row}"].alignment = center()

    if qty and price:
        ws1[f"F{row}"] = f"=C{row}*D{row}*(1-E{row}/100)" if disc else f"=C{row}*D{row}"
    else:
        ws1[f"F{row}"] = ""
    ws1[f"F{row}"].font = body_font(bold=True, size=9)
    ws1[f"F{row}"].fill = fill(bg)
    ws1[f"F{row}"].alignment = center()
    ws1[f"F{row}"].number_format = "$#,##0.00"

    for col in ["A","B","C","D","E","F"]:
        ws1[f"{col}{row}"].border = border(color=gris_borde)

# --- TOTALES ---
tot_row = 21
ws1.row_dimensions[tot_row].height = 4

sub_row = 22
ws1.row_dimensions[sub_row].height = 20
ws1.merge_cells(f"B{sub_row}:E{sub_row}")
ws1[f"B{sub_row}"] = "Subtotal"
ws1[f"B{sub_row}"].font = body_font(bold=True, size=10)
ws1[f"B{sub_row}"].alignment = Alignment(horizontal="right", vertical="center")
ws1[f"F{sub_row}"] = "=SUM(F12:F19)"
ws1[f"F{sub_row}"].font = body_font(bold=True, size=10)
ws1[f"F{sub_row}"].alignment = center()
ws1[f"F{sub_row}"].number_format = "$#,##0.00"
ws1[f"F{sub_row}"].fill = fill(gris_claro)
ws1[f"F{sub_row}"].border = border(color=gris_borde)

iva_row = 23
ws1.row_dimensions[iva_row].height = 20
ws1.merge_cells(f"B{iva_row}:E{iva_row}")
ws1[f"B{iva_row}"] = "IVA (21%)"
ws1[f"B{iva_row}"].font = body_font(size=10, color=texto_gris)
ws1[f"B{iva_row}"].alignment = Alignment(horizontal="right", vertical="center")
ws1[f"F{iva_row}"] = f"=F{sub_row}*0.21"
ws1[f"F{iva_row}"].font = body_font(size=10, color=texto_gris)
ws1[f"F{iva_row}"].alignment = center()
ws1[f"F{iva_row}"].number_format = "$#,##0.00"
ws1[f"F{iva_row}"].fill = fill(gris_claro)
ws1[f"F{iva_row}"].border = border(color=gris_borde)

total_row = 24
ws1.row_dimensions[total_row].height = 26
ws1.merge_cells(f"B{total_row}:E{total_row}")
ws1[f"B{total_row}"] = "TOTAL"
ws1[f"B{total_row}"].font = header_font(size=12, color=blanco)
ws1[f"B{total_row}"].fill = fill(azul_oscuro)
ws1[f"B{total_row}"].alignment = Alignment(horizontal="right", vertical="center", indent=2)
ws1[f"F{total_row}"] = f"=F{sub_row}+F{iva_row}"
ws1[f"F{total_row}"].font = Font(name="Arial", bold=True, size=13, color=blanco)
ws1[f"F{total_row}"].fill = fill(azul_medio)
ws1[f"F{total_row}"].alignment = center()
ws1[f"F{total_row}"].number_format = "$#,##0.00"

# --- NOTAS ---
ws1.row_dimensions[26].height = 20
ws1.merge_cells("B26:F26")
ws1["B26"] = "TÉRMINOS Y CONDICIONES"
ws1["B26"].font = header_font(size=9, color=azul_oscuro)
ws1["B26"].fill = fill(azul_claro)
ws1["B26"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

notas = [
    "· El presupuesto tiene validez de 30 días desde la fecha de emisión.",
    "· El trabajo comienza una vez confirmado el adelanto del 50%.",
    "· Las revisiones incluidas son 2 rondas por entregable.",
    "· Los plazos de entrega se cuentan desde la aprobación y pago del adelanto.",
]
for i, nota in enumerate(notas):
    r = 27 + i
    ws1.row_dimensions[r].height = 16
    ws1.merge_cells(f"B{r}:F{r}")
    ws1[f"B{r}"] = nota
    ws1[f"B{r}"].font = body_font(size=9, color=texto_gris)
    ws1[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

# Firma
ws1.row_dimensions[32].height = 20
ws1.merge_cells("B32:C33")
ws1["B32"] = "Firma del cliente"
ws1["B32"].font = body_font(size=9, color=texto_gris)
ws1["B32"].alignment = center()
ws1["B32"].border = Border(top=Side(style="medium", color=azul_oscuro))

ws1.merge_cells("E32:F33")
ws1["E32"] = "Firma Vortex Studio"
ws1["E32"].font = body_font(size=9, color=texto_gris)
ws1["E32"].alignment = center()
ws1["E32"].border = Border(top=Side(style="medium", color=azul_oscuro))

# ---- HOJA 2: SEGUIMIENTO ----
ws2 = wb.create_sheet("Seguimiento de proyectos")
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 16
ws2.column_dimensions["H"].width = 14

ws2.row_dimensions[2].height = 40
ws2.merge_cells("B2:H2")
ws2["B2"] = "SEGUIMIENTO DE PROYECTOS — VORTEX STUDIO"
ws2["B2"].font = Font(name="Arial", bold=True, size=16, color=blanco)
ws2["B2"].fill = fill(azul_oscuro)
ws2["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

ws2.row_dimensions[4].height = 22
seg_headers = ["#", "Proyecto / Cliente", "Servicio", "Inicio", "Entrega", "Estado", "Monto (USD)", "Cobrado"]
seg_cols = ["A","B","C","D","E","F","G","H"]
for col, h in zip(seg_cols, seg_headers):
    c = ws2[f"{col}4"]
    c.value = h
    c.font = header_font(size=9)
    c.fill = fill(azul_oscuro)
    c.alignment = center()
    c.border = border(color=azul_medio)

estados = {
    "En curso": ("FFF3CD", "856404"),
    "Entregado": ("D1FAE5", "065F46"),
    "Pendiente": ("FEE2E2", "991B1B"),
    "Finalizado": ("E8F0FE", "0F3A9E"),
}

seg_data = [
    (1, "Estudio Mira", "Diseño web", "01/05/2025", "08/05/2025", "Finalizado", 150, 150),
    (2, "Café Lumière", "Redes sociales", "05/05/2025", "05/06/2025", "En curso", 80, 40),
    (3, "Dr. Fernández", "Identidad de marca", "10/05/2025", "17/05/2025", "Entregado", 120, 60),
    (4, "", "", "", "", "", 0, 0),
    (5, "", "", "", "", "", 0, 0),
    (6, "", "", "", "", "", 0, 0),
    (7, "", "", "", "", "", 0, 0),
    (8, "", "", "", "", "", 0, 0),
]

for idx, (num, proj, serv, ini, fin, estado, monto, cobrado) in enumerate(seg_data):
    row = 5 + idx
    ws2.row_dimensions[row].height = 20
    bg = blanco if idx % 2 == 0 else gris_claro

    vals = [num or "", proj, serv, ini, fin, estado, monto if monto else "", cobrado if cobrado else ""]
    aligns = [center(), left(), left(), center(), center(), center(), center(), center()]
    fmts = [None, None, None, None, None, None, "$#,##0.00", "$#,##0.00"]

    for col, val, aln, fmt in zip(seg_cols, vals, aligns, fmts):
        c = ws2[f"{col}{row}"]
        c.value = val
        c.font = body_font(size=9)
        c.fill = fill(bg)
        c.alignment = aln
        c.border = border(color=gris_borde)
        if fmt and val:
            c.number_format = fmt

    if estado and estado in estados:
        bg_e, fg_e = estados[estado]
        ws2[f"F{row}"].fill = fill(bg_e)
        ws2[f"F{row}"].font = Font(name="Arial", bold=True, size=9, color=fg_e)

# Totales seguimiento
tot = 14
ws2.row_dimensions[tot].height = 22
ws2.merge_cells(f"B{tot}:F{tot}")
ws2[f"B{tot}"] = "TOTAL FACTURADO / COBRADO"
ws2[f"B{tot}"].font = header_font(size=9, color=blanco)
ws2[f"B{tot}"].fill = fill(azul_oscuro)
ws2[f"B{tot}"].alignment = Alignment(horizontal="right", vertical="center", indent=2)

ws2[f"G{tot}"] = "=SUM(G5:G12)"
ws2[f"G{tot}"].font = Font(name="Arial", bold=True, size=10, color=blanco)
ws2[f"G{tot}"].fill = fill(azul_medio)
ws2[f"G{tot}"].alignment = center()
ws2[f"G{tot}"].number_format = "$#,##0.00"

ws2[f"H{tot}"] = "=SUM(H5:H12)"
ws2[f"H{tot}"].font = Font(name="Arial", bold=True, size=10, color=blanco)
ws2[f"H{tot}"].fill = fill(azul_medio)
ws2[f"H{tot}"].alignment = center()
ws2[f"H{tot}"].number_format = "$#,##0.00"

# ---- HOJA 3: CALCULADORA DE PRECIOS ----
ws3 = wb.create_sheet("Calculadora de precios")
ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18

ws3.row_dimensions[2].height = 40
ws3.merge_cells("B2:D2")
ws3["B2"] = "CALCULADORA DE PRECIO DE SERVICIOS"
ws3["B2"].font = Font(name="Arial", bold=True, size=14, color=blanco)
ws3["B2"].fill = fill(azul_oscuro)
ws3["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

ws3.row_dimensions[4].height = 22
ws3.merge_cells("B4:D4")
ws3["B4"] = "INPUTS — Completá los campos en azul"
ws3["B4"].font = header_font(size=9, color=azul_oscuro)
ws3["B4"].fill = fill(azul_claro)
ws3["B4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

inputs = [
    ("Horas estimadas del proyecto", 10, "hs"),
    ("Valor hora deseado (USD)", 25, "USD/h"),
    ("Costo de herramientas / licencias", 20, "USD"),
    ("Margen de ganancia (%)", 30, "%"),
    ("Descuento a aplicar (%)", 0, "%"),
]

for idx, (label, val, unit) in enumerate(inputs):
    row = 5 + idx
    ws3.row_dimensions[row].height = 20
    ws3[f"B{row}"] = label
    ws3[f"B{row}"].font = body_font(size=10)
    ws3[f"B{row}"].alignment = left()
    ws3[f"C{row}"] = val
    ws3[f"C{row}"].font = Font(name="Arial", bold=True, size=10, color="0000FF")
    ws3[f"C{row}"].alignment = center()
    ws3[f"C{row}"].fill = fill("EEF2FF")
    ws3[f"C{row}"].border = Border(bottom=Side(style="medium", color=azul_medio))
    ws3[f"D{row}"] = unit
    ws3[f"D{row}"].font = body_font(size=9, color=texto_gris)
    ws3[f"D{row}"].alignment = center()

ws3.row_dimensions[11].height = 22
ws3.merge_cells("B11:D11")
ws3["B11"] = "RESULTADO"
ws3["B11"].font = header_font(size=9, color=azul_oscuro)
ws3["B11"].fill = fill(azul_claro)
ws3["B11"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

resultados = [
    ("Costo base", "=C5*C6+C7"),
    ("Costo con margen", "=B13*(1+C8/100)"),
    ("Descuento", "=B14*(C9/100)"),
    ("PRECIO FINAL (USD)", "=B14-B15"),
]

for idx, (label, formula) in enumerate(resultados):
    row = 12 + idx
    ws3.row_dimensions[row].height = 22
    ws3[f"B{row}"] = label
    is_total = label.startswith("PRECIO")
    ws3[f"B{row}"].font = header_font(size=10, color=blanco) if is_total else body_font(bold=is_total, size=10)
    ws3[f"B{row}"].fill = fill(azul_oscuro) if is_total else fill(gris_claro)
    ws3[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws3[f"C{row}"] = formula
    ws3[f"C{row}"].font = Font(name="Arial", bold=True, size=11 if is_total else 10, color=blanco if is_total else azul_oscuro)
    ws3[f"C{row}"].fill = fill(azul_medio) if is_total else fill(gris_claro)
    ws3[f"C{row}"].alignment = center()
    ws3[f"C{row}"].number_format = "$#,##0.00"
    ws3[f"C{row}"].border = border(color=gris_borde)

wb.save("/home/claude/productos-digitales/Pack-Presupuestos-Pro-VortexStudio.xlsx")
print("Excel creado OK")
